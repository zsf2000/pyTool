import openpyxl
import pandas as pd
import os
import xlwings as xw
from openpyxl.utils import get_column_letter
# 创建一个新的工作簿
workbook = openpyxl.load_workbook('L2ods数据.xlsx')

# 选择默认的工作表
worksheet = workbook.active

# 设置要更改的列范围
column_range = range(1, 6)

# 设置要应用的宽度
column_width = 80

# 遍历列范围并设置每个列的宽度
for column in column_range:
    column_letter = get_column_letter(column)
    # column_letter = worksheet.c
    worksheet.column_dimensions[column_letter].width = column_width

# 保存工作簿
workbook.save("example.xlsx")
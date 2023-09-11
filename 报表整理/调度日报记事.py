# -*- coding:utf-8 -*-

from openpyxl import load_workbook
import pandas as pd
import os
import xlwings as xw


pd.set_option('display.max_columns',1000)
pd.set_option('display.width',1000)

workbook = load_workbook(r'D:\生产管理\调度日报模板\L3导出.xlsx')
#D:\生产管理\调度日报模板\L3导出.xlsx
print(workbook.sheetnames)

sheet=workbook['Sheet1']
print(sheet)
print(sheet.dimensions)
js_data=sheet['H2']
#sheet['H2']=pd.to_datetime(sheet['H2'])
js_data=sheet['H2']
print(js_data.value)
js_yl=sheet['A33']
js_yl1=sheet['D33']
print(js_yl.value,js_yl1.value,'天')
js_yl2=sheet['A34']
print(js_yl2.value)

js_dl=sheet['N33']
js_dl1=sheet['t33']
print(js_dl.value,js_dl1.value,'天')
js_dl2=sheet['N34']
js_dl21=sheet['O34']
print(js_dl2.value,js_dl21.value)
js_dl3=sheet['N40']
js_dl31=sheet['O40']
print(js_dl3.value,js_dl31.value)
js_dl4=sheet['N47']
js_dl41=sheet['O47']
print(js_dl4.value,js_dl41.value)

js_co=sheet['x88']
js_co_1=sheet['AC88']
print(js_co.value,js_co_1.value,'天')
js_co1=sheet['x89']
js_co11=sheet['z89']
print(js_co1.value,js_co11.value)
js_co3=sheet['x93']
js_co31=sheet['z93']
print(js_co3.value,js_co31.value)
js_co4=sheet['x97']
js_co41=sheet['z97']
print(js_dl4.value,js_dl41.value)

js_co5=sheet['x101']
js_co51=sheet['z101']
print(js_co5.value,js_co51.value)

js_bf=sheet['A88']
js_bf_1=sheet['D88']
print(js_co.value,js_co_1.value,'天')
js_bf1=sheet['A89']
js_bf11=sheet['B89']
print(js_bf1.value,js_bf11.value)
js_bf2=sheet['k89']
js_bf21=sheet['L89']
print(js_bf2.value,js_bf21.value)
js_bf3=sheet['A94']
js_bf31=sheet['B94']
print(js_bf3.value,js_bf31.value)
js_bf4=sheet['k94']
js_bf41=sheet['L94']
print(js_bf4.value,js_bf41.value)

js_bf5=sheet['A100']
print(js_bf5.value)

js= pd.DataFrame(columns=['日期','原料','煤焦','2DL','3DL','4DL','1CO','3CO','4CO','1BF','2BF','3BF','4BF','BF'])
js.loc[0]=[js_data.value,js_yl2.value,js_co51.value,js_dl21.value,js_dl31.value,js_dl41.value,js_co11.value,js_co31.value,js_co41.value,js_bf11.value,js_bf21.value,js_bf31.value,js_bf41.value,js_bf5.value,]
print(js)



js0=pd.read_excel(r'D:\生产管理\调度日报记事.xlsx')

print(js0)
js0=pd.concat([js0,js])


js0.to_excel(excel_writer=r'D:\生产管理\调度日报记事.xlsx',sheet_name='调度日报记事',index=False,columns=['日期','原料','煤焦','2DL','3DL','4DL','1CO','3CO','4CO','1BF','2BF','3BF','4BF','BF'])

